#!/usr/bin/env python3
"""
Enchant Materials Generator

Reads enchant.xlsx and generates:
1. reforged/specs/enchant-materials.yaml - MaterialEnchantData definitions
2. reforged/specs/enchant-item-links.yaml - Items updateWhere rules

Uses DSL parameterized definitions ($extends, $with, $params) for compact output.

ID Scheme for materialEnchantId:
  Base: 20000
  Pattern: 20000 + (tierIndex * 10000) + (levelRangeIndex * 1000) + (slotGroupIndex * 100) + rank

  tierIndex: 0=default (+12), 1=mythic (+15)
  levelRangeIndex: 0=1-37, 1=38-49, 2=50-57, 3=58+
  slotGroupIndex: 0=WeC (Weapons & Chest), 1=GeB (Gloves & Boots)
"""

import argparse
import sys
from dataclasses import dataclass
from itertools import groupby
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# Configuration
ALKAHEST_ID = 21351
FEEDSTOCK_BASE_ID = 94100  # Feedstock ID = 94100 + rank

LEVEL_RANGES = [
    (0, "1..37", 1, 37),
    (1, "38..49", 38, 49),
    (2, "50..57", 50, 57),
    (3, "58..65", 58, 65),  # 58+ represented as 58..65
]

SLOT_GROUPS = [
    (0, "WeC", ["EQUIP_WEAPON", "EQUIP_ARMOR_BODY"]),
    (1, "GeB", ["EQUIP_ARMOR_ARM", "EQUIP_ARMOR_LEG"]),
]

# Rank distribution per level range (from analysis)
RANKS_BY_LEVEL_RANGE = {
    0: [1, 2],           # Level 1-37
    1: [2],              # Level 38-49
    2: [2, 3],           # Level 50-57
    3: list(range(3, 23)),  # Level 58+: ranks 3-22
}

# Enchant tiers: (tier_idx, max_enchant_count, grade_filter)
# rareGrade is a string attribute — use list syntax instead of range
ENCHANT_TIERS = [
    (0, 12, "[Common, Uncommon, Rare, Superior]"),  # default: grades 0-3
    (1, 15, "Mythic"),                               # mythic: grade 4
]


@dataclass
class EnchantStep:
    step: int
    prob: float
    alkahest_amount: int
    feedstock_amount: int


@dataclass
class MaterialEnchantConfig:
    material_enchant_id: int
    tier_idx: int
    max_enchant_count: int
    grade_filter: str
    level_range_idx: int
    level_range_str: str
    slot_group_idx: int
    slot_group_name: str
    combat_item_types: list
    rank: int
    steps: list  # List of EnchantStep


def calculate_material_enchant_id(tier_idx: int, level_range_idx: int, slot_group_idx: int, rank: int) -> int:
    """Calculate materialEnchantId using the defined scheme."""
    return 20000 + (tier_idx * 10000) + (level_range_idx * 1000) + (slot_group_idx * 100) + rank


def read_excel_sheet(wb, sheet_name: str) -> dict:
    """Read a sheet and return data as dict[enchant_step][level_range_col] = value."""
    ws = wb[sheet_name]

    # Column headers are in row 1: first col is enchant step, rest are level ranges
    col_headers = []
    for col in range(2, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header:
            col_headers.append((col, str(header).strip()))

    data = {}
    for row in range(2, ws.max_row + 1):
        step_val = ws.cell(row=row, column=1).value
        if step_val is None:
            continue
        step = int(step_val)
        data[step] = {}
        for col, header in col_headers:
            val = ws.cell(row=row, column=col).value
            if val is not None:
                data[step][header] = float(val) if isinstance(val, (int, float)) else val

    return data


def map_level_range_to_col(level_range_idx: int) -> str:
    """Map level range index to Excel column header."""
    col_map = {
        0: "1 ~37",
        1: "38 ~49",
        2: "50 ~57",
        3: "58",
    }
    return col_map[level_range_idx]


def load_enchant_data(xlsx_path: str) -> tuple:
    """Load all enchant data from Excel file."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    chance_data = read_excel_sheet(wb, "Chance")
    alkahest_wec = read_excel_sheet(wb, "Alkahest WeC")
    alkahest_geb = read_excel_sheet(wb, "Alkahest GeB")
    feedstock_wec = read_excel_sheet(wb, "Feedstock WeC")
    feedstock_geb = read_excel_sheet(wb, "Feedstock GeB")

    wb.close()

    return chance_data, alkahest_wec, alkahest_geb, feedstock_wec, feedstock_geb


def build_material_enchant_configs(
    chance_data: dict,
    alkahest_wec: dict,
    alkahest_geb: dict,
    feedstock_wec: dict,
    feedstock_geb: dict,
) -> list:
    """Build all MaterialEnchantConfig objects."""
    configs = []

    for tier_idx, max_enchant_count, grade_filter in ENCHANT_TIERS:
        for level_range_idx, level_range_str, level_min, level_max in LEVEL_RANGES:
            col_name = map_level_range_to_col(level_range_idx)
            ranks = RANKS_BY_LEVEL_RANGE[level_range_idx]

            for slot_group_idx, slot_group_name, combat_types in SLOT_GROUPS:
                # Select alkahest/feedstock data based on slot group
                alkahest_data = alkahest_wec if slot_group_idx == 0 else alkahest_geb
                feedstock_data = feedstock_wec if slot_group_idx == 0 else feedstock_geb

                for rank in ranks:
                    material_enchant_id = calculate_material_enchant_id(
                        tier_idx, level_range_idx, slot_group_idx, rank
                    )

                    steps = []
                    for step in range(max_enchant_count):
                        prob = chance_data.get(step, {}).get(col_name, 1.0)
                        alkahest_amount = int(alkahest_data.get(step, {}).get(col_name, 0))
                        feedstock_amount = int(feedstock_data.get(step, {}).get(col_name, 0))

                        steps.append(EnchantStep(
                            step=step,
                            prob=prob,
                            alkahest_amount=alkahest_amount,
                            feedstock_amount=feedstock_amount,
                        ))

                    configs.append(MaterialEnchantConfig(
                        material_enchant_id=material_enchant_id,
                        tier_idx=tier_idx,
                        max_enchant_count=max_enchant_count,
                        grade_filter=grade_filter,
                        level_range_idx=level_range_idx,
                        level_range_str=level_range_str,
                        slot_group_idx=slot_group_idx,
                        slot_group_name=slot_group_name,
                        combat_item_types=combat_types,
                        rank=rank,
                        steps=steps,
                    ))

    return configs


def format_prob(p: float) -> str:
    """Format probability value."""
    if p == 1.0:
        return "1.0"
    elif p == 0.0:
        return "0.0"
    else:
        return f"{p:.2f}".rstrip('0').rstrip('.')


_TIER_NAMES = {0: "Default", 1: "Mythic"}
_LEVEL_RANGE_NAMES = {0: "L1_37", 1: "L38_49", 2: "L50_57", 3: "L58"}
_SLOT_GROUP_NAMES = {0: "WeC", 1: "GeB"}


def _definition_name(tier_idx: int, level_range_idx: int, slot_group_idx: int) -> str:
    """Compute definition name from group indices."""
    return f"{_TIER_NAMES[tier_idx]}_{_LEVEL_RANGE_NAMES[level_range_idx]}_{_SLOT_GROUP_NAMES[slot_group_idx]}"


def _collect_material_tiers(configs: list) -> dict:
    """Collect unique (alkahest, feedstock) amount pairs per slot group, in order of first appearance."""
    tiers = {0: [], 1: []}
    seen = {0: set(), 1: set()}
    for config in configs:
        sg = config.slot_group_idx
        for step in config.steps:
            key = (step.alkahest_amount, step.feedstock_amount)
            if key not in seen[sg]:
                seen[sg].add(key)
                tiers[sg].append(key)
    return tiers


def _step_mat_name(slot_group_idx: int, tier_index: int) -> str:
    """Get step material definition name."""
    return f"_{_SLOT_GROUP_NAMES[slot_group_idx]}_Mat_{chr(ord('A') + tier_index)}"


def _build_step_mat_lookup(tiers: dict) -> dict:
    """Build (slot_group_idx, alkahest, feedstock) -> definition name lookup."""
    lookup = {}
    for sg_idx, pairs in tiers.items():
        for i, (alk, feed) in enumerate(pairs):
            lookup[(sg_idx, alk, feed)] = _step_mat_name(sg_idx, i)
    return lookup


def _emit_step_material_definitions(lines: list, tiers: dict) -> None:
    """Emit step material definitions for each slot group."""
    for sg_idx in (0, 1):
        sg_name = _SLOT_GROUP_NAMES[sg_idx]
        lines.append(f"  # \u2500\u2500 {sg_name} material tiers \u2500\u2500")
        for i, (alk, feed) in enumerate(tiers[sg_idx]):
            name = _step_mat_name(sg_idx, i)
            lines.append(f"  {name}:")
            lines.append("    requiredMoney: 0")
            lines.append("    materials:")
            lines.append(f"      - id: {ALKAHEST_ID}")
            lines.append("        type: Item")
            lines.append(f"        amount: {alk}")
            lines.append("      - id: $FEEDSTOCK_ID")
            lines.append("        type: Item")
            lines.append(f"        amount: {feed}")
            lines.append("")


def _emit_definition(lines: list, name: str, max_steps: int, steps: list,
                     slot_group_idx: int, step_mat_lookup: dict) -> None:
    """Emit a parameterized definition using step material $extends."""
    lines.append(f"  {name}:")
    lines.append("    $params: [ENCHANT_ID, FEEDSTOCK_ID]")
    lines.append("    materialEnchantId: $ENCHANT_ID")
    lines.append(f"    maxEnchantCount: {max_steps}")
    lines.append("    materialItems:")
    for step in steps:
        mat_def = step_mat_lookup[(slot_group_idx, step.alkahest_amount, step.feedstock_amount)]
        lines.append(f"      - $extends: {mat_def}")
        lines.append(f"        enchantStep: {step.step}")
        lines.append(f"        enchantProb: {format_prob(step.prob)}")


def _group_key(config: MaterialEnchantConfig) -> tuple:
    """Grouping key for configs: (tier_idx, level_range_idx, slot_group_idx)."""
    return (config.tier_idx, config.level_range_idx, config.slot_group_idx)


def _emit_entries(lines: list, group: list, def_name: str) -> None:
    """Emit compact $extends + $with entries for a group of configs."""
    first = group[0]
    tier_label = "mythic" if first.tier_idx == 1 else "default"
    step_count = first.max_enchant_count

    if len(group) == 1:
        rank_desc = f"Rank {first.rank}"
    else:
        rank_desc = f"Ranks {group[0].rank}..{group[-1].rank}"

    if first.tier_idx == 0:
        lines.append(f"    # \u2500\u2500 {first.slot_group_name}, Level {first.level_range_str}, {rank_desc} \u2500\u2500")
    else:
        lines.append(f"    # \u2500\u2500 {first.slot_group_name}, Level {first.level_range_str}, {rank_desc} ({tier_label} {step_count}-step) \u2500\u2500")

    for config in group:
        feedstock_id = FEEDSTOCK_BASE_ID + config.rank
        lines.append(f"    - $extends: {def_name}")
        lines.append(f"      $with: {{ ENCHANT_ID: {config.material_enchant_id}, FEEDSTOCK_ID: {feedstock_id} }}")


def generate_material_enchants_yaml(configs: list) -> str:
    """Generate the materialEnchants YAML spec using parameterized definitions."""
    lines = [
        "# Enchant Materials System - MaterialEnchantData",
        "# Auto-generated by generate_enchant_materials.py",
        "# Uses DSL parameterized definitions ($extends, $with, $params)",
        "",
        "spec:",
        '  version: "1.0"',
        "  schema: v92",
        "",
        "definitions:",
    ]

    # Collect material tiers and build lookup
    tiers = _collect_material_tiers(configs)
    step_mat_lookup = _build_step_mat_lookup(tiers)

    # Emit step material definitions
    _emit_step_material_definitions(lines, tiers)

    # Collect groups and emit entry definitions
    groups = []
    for key, group_iter in groupby(configs, key=_group_key):
        groups.append((key, list(group_iter)))

    for (tier_idx, level_range_idx, slot_group_idx), group in groups:
        def_name = _definition_name(tier_idx, level_range_idx, slot_group_idx)
        representative = group[0]
        _emit_definition(lines, def_name, representative.max_enchant_count,
                         representative.steps, slot_group_idx, step_mat_lookup)
        lines.append("")

    # Emit upsert entries
    lines.append("materialEnchants:")
    lines.append("  upsert:")

    for (tier_idx, level_range_idx, slot_group_idx), group in groups:
        def_name = _definition_name(tier_idx, level_range_idx, slot_group_idx)
        _emit_entries(lines, group, def_name)
        lines.append("")

    return "\n".join(lines)


def generate_item_links_yaml(configs: list) -> str:
    """Generate the items updateWhere YAML spec."""
    lines = [
        "# Enchant Materials System - Item Links",
        "# Auto-generated by generate_enchant_materials.py",
        "",
        "spec:",
        '  version: "1.0"',
        "  schema: v92",
        "",
        "items:",
        "  updateWhere:",
    ]

    for config in configs:
        combat_types_yaml = "[" + ", ".join(config.combat_item_types) + "]"

        tier_label = "mythic" if config.tier_idx == 1 else "default"
        lines.append(f"    # {config.slot_group_name}, Level {config.level_range_str}, Rank {config.rank} ({tier_label})")
        lines.append("    - filter:")
        lines.append("        enchantEnable: true")
        lines.append(f"        combatItemType: {combat_types_yaml}")
        lines.append(f"        level: {config.level_range_str}")
        lines.append(f"        rank: {config.rank}")
        lines.append(f"        rareGrade: {config.grade_filter}")
        lines.append("      changes:")
        lines.append(f"        linkMaterialEnchantId: {config.material_enchant_id}")
        lines.append("")

    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser(description="Generate enchant materials YAML specs")
    parser.add_argument("--patch", help="Patch folder name (e.g. 001). Output goes to reforged/specs/patches/{patch}/")
    args = parser.parse_args()

    # Determine paths
    script_dir = Path(__file__).parent
    xlsx_path = script_dir.parent.parent / "data" / "enchant.xlsx"
    project_root = script_dir.parent.parent.parent  # reforged-server-content

    if args.patch:
        specs_dir = project_root / "reforged" / "specs" / "patches" / args.patch
        specs_dir.mkdir(parents=True, exist_ok=True)
    else:
        specs_dir = project_root / "reforged" / "specs"

    if not xlsx_path.exists():
        print(f"Error: {xlsx_path} not found")
        sys.exit(1)

    print(f"Reading {xlsx_path}...")
    chance_data, alkahest_wec, alkahest_geb, feedstock_wec, feedstock_geb = load_enchant_data(str(xlsx_path))

    print("Building material enchant configurations...")
    configs = build_material_enchant_configs(
        chance_data, alkahest_wec, alkahest_geb, feedstock_wec, feedstock_geb
    )
    print(f"Generated {len(configs)} configurations")

    # Generate YAML specs
    materials_yaml = generate_material_enchants_yaml(configs)
    item_links_yaml = generate_item_links_yaml(configs)

    # Write output files
    materials_path = specs_dir / "04-enchant-materials.yaml"
    item_links_path = specs_dir / "05-enchant-item-links.yaml"

    print(f"Writing {materials_path}...")
    with open(materials_path, "w", encoding="utf-8") as f:
        f.write(materials_yaml)

    print(f"Writing {item_links_path}...")
    with open(item_links_path, "w", encoding="utf-8") as f:
        f.write(item_links_yaml)

    # Count lines for comparison
    old_line_estimate = len(configs) * 150  # ~150 lines per entry in old format
    new_lines = len(materials_yaml.split('\n'))

    print(f"\nDone!")
    print(f"  Old format estimate: ~{old_line_estimate} lines")
    print(f"  New format: {new_lines} lines")
    print(f"  Reduction: ~{100 - (new_lines * 100 // old_line_estimate)}%")
    print(f"\nNext steps:")
    print(f"  dsl validate \"{materials_path.relative_to(project_root)}\" --path \"D:\\dev\\mmogate\\tera92\\server\\Datasheet\"")
    print(f"  dsl apply \"{materials_path.relative_to(project_root)}\" --path \"D:\\dev\\mmogate\\tera92\\server\\Datasheet\"")


if __name__ == "__main__":
    main()
